import openpyxl as op
import pprint as pp

subcategories_dict = {}


filename = 'Бланк заказа.xlsx'
#Открываем на чтание файл
wb = op.load_workbook(filename, data_only=True)
#Выбираем активный лист файла
sheet = wb.active
#Находим последнюю строку с данными
max_rows = sheet.max_row
#Читаем значение ячейки
#print(sheet.cell(row=7, column=2).value)
#Пробигаем по всем строкам
for i in range(7,max_rows+1):
    sku = sheet.cell(row=i, column=2).value
    subcategory = sheet.cell(row=i, column=12).value
    if not sku:
        continue
#Формируем перечень артикулов и подкатегорий
    if subcategory not in subcategories_dict:
        subcategories_dict[subcategory] = [sku]
    else:
        subcategories_dict[subcategory].append(sku)
#Читаемый вывод в консоль
#pp.pprint(subcategories_dict)
#Сортировка словаря
sortedict = dict(sorted(subcategories_dict.items()))

#Записываем в файл
with open("subcategories.ini", 'w') as myfile:

    for key, values in sortedict.items():
        string_values = ', '.join(values)
        string_to_write = key + ' = ' + string_values + '\n'
        myfile.write(string_to_write)


